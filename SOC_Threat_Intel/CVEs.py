import requests
import os
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.chart import PieChart, BarChart, Reference
from collections import Counter
from tqdm import tqdm
from deep_translator import GoogleTranslator
import time
import json

# API
load_dotenv()
VULNCHECK_TOKEN = os.getenv("VULNCHECK_TOKEN")
if not VULNCHECK_TOKEN:
    print("[!] Error: No se encontró el token de VulnCheck\nAsegurate de colocar el token en tu archivo .env, o en caso de no tenerlo, crea su respectivo .env")
    exit()

# Rutas
DIRECTORIO_ACTUAL = os.path.dirname(os.path.abspath(__file__))
RUTA_REPO_LOCAL = os.path.join(DIRECTORIO_ACTUAL, "cvelistV5")
ARCHIVO_TXT = os.path.join(DIRECTORIO_ACTUAL, "lista_cves.txt")
ARCHIVO_EXCEL = os.path.join(DIRECTORIO_ACTUAL, "bitacora_vulnerabilidades.xlsx")

MAPA_SEVERIDAD = {
    "CRITICAL": "Crítica", "HIGH": "Alta", "MEDIUM": "Media", "LOW": "Baja", "N/A": "No evaluado"
}
MAPA_VECTOR = {
    "NETWORK": "Red (Remoto)", "ADJACENT_NETWORK": "Red Adyacente", "LOCAL": "Local", 
    "PHYSICAL": "Físico", "N/A": "Desconocido"
}

MAPA_TACTICO_AVANZADO = {
    "89": "ARQUITECTURA: Migrar a ORM o Prepared Statements. RED: Desplegar reglas WAF contra inyecciones SQL. SISTEMA: Aplicar Principio de Menor Privilegio (PoLP).",
    "78": "ARQUITECTURA: Sustituir llamadas al SO (system/exec) por APIs nativas. SISTEMA: Ejecutar el aplicativo en un entorno 'chroot' o contenedor aislado.",
    "79": "ARQUITECTURA: Forzar Context-Aware Output Encoding. RED: Configurar cabeceras CSP restrictivas y aplicar la bandera HttpOnly en cookies.",
    "416": "SISTEMA: Auditar código con AddressSanitizer (ASan). Evitar punteros colgantes anulándolos tras la liberación.",
    "119": "SISTEMA: Habilitar protecciones (ASLR, DEP/NX). ARQUITECTURA: Validar límites de buffers de forma estricta.",
    "125": "SISTEMA: Aislar el proceso vulnerable. ARQUITECTURA: Implementar comprobaciones estrictas de límites de arreglos (Bounds Checking).",
    "22": "ARQUITECTURA: No exponer nombres de archivos. SISTEMA: Restringir acceso al File System usando ACLs estrictas y chroot jails.",
    "502": "ARQUITECTURA: Prohibir la deserialización de datos no confiables. Migrar a formatos puros (JSON/XML).",
    "352": "ARQUITECTURA: Implementar tokens Anti-CSRF únicos. Configurar atributo SameSite=Lax/Strict en cookies de sesión.",
    "918": "RED: Bloquear la resolución de IPs internas (127.0.0.1, 169.254.x.x). ARQUITECTURA: Implementar un Allowlist estricto de URLs."
}

CACHE_CWE = {}
traductor = GoogleTranslator(source='en', target='es')

def traducir_texto(texto):
    if not texto or texto == "Sin descripción": return texto
    try: return traductor.translate(texto)
    except Exception: return f"[Error traducción] {texto}"

def generar_recomendacion_hibrida(cwe_string, vector_en, solucion_basica):
    try: cwe_num = str(cwe_string).split("-")[1].split()[0]
    except IndexError: cwe_num = ""

    if cwe_num in MAPA_TACTICO_AVANZADO:
        return f"{MAPA_TACTICO_AVANZADO[cwe_num]}\n{solucion_basica}"

    if cwe_num in CACHE_CWE:
        return f"{CACHE_CWE[cwe_num]}\n{solucion_basica}"

    if cwe_num:
        url = f"https://cwe-api.mitre.org/api/v1/cwe/{cwe_num}"
        try:
            time.sleep(0.5) 
            response = requests.get(url, timeout=10)
            if response.status_code == 200:
                mitigaciones = response.json().get("Weaknesses", [])[0].get("Potential_Mitigations", [])
                if mitigaciones:
                    mitigacion_seleccionada = mitigaciones[0]
                    for mit in mitigaciones:
                        fases = mit.get("Phase", [])
                        if "Operation" in fases or "Implementation" in fases:
                            mitigacion_seleccionada = mit
                            break

                    fase = mitigacion_seleccionada.get("Phase", ["Desconocida"])[0]
                    desc_en = str(mitigacion_seleccionada.get("Description", "")).replace('\n', ' ').strip()
                    if desc_en:
                        desc_es = traducir_texto(desc_en)
                        fase_es = traducir_texto(fase)
                        # Se respeta el formato sin corchetes y tal como viene el input
                        mitigacion_oficial = f"Fase {fase_es}: {desc_es}"
                        CACHE_CWE[cwe_num] = mitigacion_oficial
                        return f"{mitigacion_oficial}\n{solucion_basica}"
        except Exception:
            pass

    if vector_en == "NETWORK":
        tactica = "Segmentar red, bloquear puertos expuestos perimetralmente (Firewall/ACLs) y aplicar firmas restrictivas en el WAF/IPS."
    elif vector_en == "LOCAL":
        tactica = "Auditar privilegios de usuarios (PoLP), restringir ejecución de binarios sospechosos y aislar el host."
    elif vector_en == "ADJACENT_NETWORK":
        tactica = "Aislar VLAN afectada, requerir autenticación fuerte 802.1X y monitorizar tráfico lateral."
    else:
        tactica = "Aplicar controles de acceso estrictos, monitoreo de logs y aislamiento preventivo."

    return f"CONTENCIÓN TÁCTICA: {tactica}\n{solucion_basica}"

def reglas_de_orden(fila):
    severidad_str = str(fila[4]).lower()
    if "crítica" in severidad_str: orden_sev = 1
    elif "alta" in severidad_str: orden_sev = 2
    elif "media" in severidad_str: orden_sev = 3
    elif "baja" in severidad_str: orden_sev = 4
    else: orden_sev = 5
    
    fecha_str = str(fila[0])
    if fecha_str in ["Desconocida", "No encontrado", "Error", "None"]:
        fecha_str = "9999-99-99" 
        
    fabricante_str = str(fila[1]).lower()
    return (orden_sev, fecha_str, fabricante_str)

def obtener_fab_tec_local(cve_id, ruta_repo):
    """Extrae Fabricante y Tecnología directamente de cvelistV5 (cve.org)"""
    fab, tec = "Variado", "Específica del CVE"
    try:
        partes = cve_id.split("-")
        if len(partes) == 3:
            anio, secuencia = partes[1], partes[2]
            carpeta_xxx = secuencia[:-3] + "xxx" if len(secuencia) >= 4 else "0xxx"
            ruta_cve = os.path.join(ruta_repo, "cves", anio, carpeta_xxx, f"{cve_id}.json")
            
            if os.path.exists(ruta_cve):
                with open(ruta_cve, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    # cve.org guarda esto en la sección "containers -> cna -> affected"
                    afectados = data.get("containers", {}).get("cna", {}).get("affected", [])
                    if afectados:
                        v = afectados[0].get("vendor", "Variado")
                        t = afectados[0].get("product", "Específica del CVE")
                        
                        # Limpieza por si el CNA reporta "n/a"
                        if v and str(v).lower() not in ["n/a", "unknown", "none"]: 
                            fab = str(v)
                        if t and str(t).lower() not in ["n/a", "unknown", "none"]: 
                            tec = str(t)
    except Exception:
        pass
    return fab, tec


print("=== BITÁCORA CVEs ===")

if not os.path.exists(ARCHIVO_TXT):
    print(f"[!] Error: No se encontró '{ARCHIVO_TXT}'.")
    exit()

with open(ARCHIVO_TXT, "r") as f:
    cves_a_buscar = [linea.strip().upper() for linea in f if linea.strip().startswith("CVE-")]

if not cves_a_buscar:
    print(f"[-] No hay identificadores válidos en '{ARCHIVO_TXT}'.")
    exit()

# Diccionario maestro para gestionar Inserciones y Actualizaciones
cves_maestro = {}

if os.path.exists(ARCHIVO_EXCEL):
    try:
        wb = openpyxl.load_workbook(ARCHIVO_EXCEL)
        if "Bitácora CVEs" in wb.sheetnames:
            ws = wb["Bitácora CVEs"]
        else:
            ws = wb.active
            ws.title = "Bitácora CVEs"
            
        # Cargamos todo el Excel existente a la memoria
        for row in range(2, ws.max_row + 1):
            fila_datos = [ws.cell(row=row, column=c).value for c in range(1, 12)]
            cve_id_existente = str(fila_datos[2]).upper()
            if any(fila_datos) and cve_id_existente.startswith("CVE-"):
                cves_maestro[cve_id_existente] = fila_datos
                
        print(f"[*] Excel detectado ({len(cves_maestro)} CVEs cargados).")
    except PermissionError:
        print(f"\n[!] ERROR: El archivo '{ARCHIVO_EXCEL}' está abierto. Ciérralo.")
        exit()
else:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bitácora CVEs"

ws.views.sheetView[0].showGridLines = True
font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
fill_header = PatternFill(start_color="2F3E46", end_color="2F3E46", fill_type="solid")
thin_border = Border(left=Side(style='thin', color='DCDDE1'), right=Side(style='thin', color='DCDDE1'),
                     top=Side(style='thin', color='DCDDE1'), bottom=Side(style='thin', color='DCDDE1'))

encabezados = [
    "Fecha Publicación", "Fabricante", "CVE Asociado", "Identificación", 
    "Severidad", "Vector", "Tecnología", "PoC/Exploit", 
    "Explotación Activa", "Recomendación / Solución", "Resumen / Descripción"
]

if ws.max_row <= 1:
    for col_idx, texto in enumerate(encabezados, 1):
        celda = ws.cell(row=1, column=col_idx, value=texto)
        celda.font = font_header
        celda.fill = fill_header
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda.border = thin_border
    ws.row_dimensions[1].height = 28


# Ya no filtramos los repetidos, los procesamos todos para buscar actualizaciones
headers_vulncheck = {"accept": "application/json", "authorization": f"Bearer {VULNCHECK_TOKEN}"}

print(f"[*] Procesando {len(cves_a_buscar)} CVE(s) en cola...\n")
for cve_id in tqdm(cves_a_buscar, desc="Consultando CVEs", unit="cve"):
    url = f"https://api.vulncheck.com/v3/index/nist-nvd2?cve={cve_id}"
    try:
        response = requests.get(url, headers=headers_vulncheck, timeout=10)
        if response.status_code == 200:
            vulns = response.json().get("data", [])
            if vulns:
                cve_data = vulns[0].get("cve", vulns[0])
                vuln_status = cve_data.get("vulnStatus", "").upper()
                if vuln_status == "REJECTED": continue
                
                fecha_pub_raw = cve_data.get("published", "")
                fecha_pub = fecha_pub_raw.split("T")[0] if "T" in fecha_pub_raw else "Desconocida"
                
                resumen_en = "Sin descripción"
                for desc in cve_data.get("descriptions", []):
                    if desc.get("lang") == "en":
                        resumen_en = desc.get("value")
                        if "** REJECT **" in resumen_en: vuln_status = "REJECTED"
                        break
                
                if vuln_status == "REJECTED": continue
                resumen_es = traducir_texto(resumen_en)
                        
                metrics = cve_data.get("metrics", {})
                severidad_en, score, vector_en = "N/A", "N/A", "N/A"
                if "cvssMetricV31" in metrics:
                    cvss_data = metrics["cvssMetricV31"][0].get("cvssData", {})
                    severidad_en = cvss_data.get('baseSeverity', 'N/A')
                    score = cvss_data.get('baseScore', 'N/A')
                    vector_en = cvss_data.get("attackVector", "N/A")
                elif "cvssMetricV30" in metrics:
                    cvss_data = metrics["cvssMetricV30"][0].get("cvssData", {})
                    severidad_en = cvss_data.get('baseSeverity', 'N/A')
                    score = cvss_data.get('baseScore', 'N/A')
                    vector_en = cvss_data.get("attackVector", "N/A")
                
                severidad_es = f"{MAPA_SEVERIDAD.get(severidad_en, severidad_en)} ({score})"
                vector_es = MAPA_VECTOR.get(vector_en, vector_en)
                    
                debilidades = cve_data.get("weaknesses", [])
                cwe = debilidades[0].get("description", [])[0].get("value", "Desconocido") if debilidades and debilidades[0].get("description") else "Desconocido"

                fabricante_raw, tecnologia_raw = obtener_fab_tec_local(cve_id, RUTA_REPO_LOCAL)
                
                if fabricante_raw == "Variado":
                    nodos = cve_data.get("configurations", [])
                    if nodos:
                        for nodo in nodos:
                            cpe_matches = nodo.get("nodes", [])[0].get("cpeMatch", [])
                            if cpe_matches:
                                partes = cpe_matches[0].get("criteria", "").split(":")
                                if len(partes) >= 5:
                                    fabricante_raw, tecnologia_raw = partes[3], partes[4]
                                break
                
                fabricante = str(fabricante_raw).capitalize()
                tecnologia = str(tecnologia_raw).capitalize()

                poc_disponible, parche_disponible = "No", False
                for ref in cve_data.get("references", []):
                    tags = ref.get("tags", [])
                    url_ref = ref.get("url", "").lower()
                    if "Exploit" in tags or "exploit-db.com" in url_ref or "github.com" in url_ref: poc_disponible = "Sí"
                    if "Patch" in tags or "Vendor Advisory" in tags: parche_disponible = True

                explotacion_activa = "Sí" if "cisaExploitAdd" in cve_data else "No"

                if "cisaRequiredAction" in cve_data:
                    solucion_base = traducir_texto(cve_data["cisaRequiredAction"])
                elif parche_disponible:
                    solucion_base = "Parche/Actualización disponible. Aplicar a la brevedad."
                else:
                    solucion_base = "Monitorear advisories del fabricante."

                solucion_final = generar_recomendacion_hibrida(cwe, vector_en, solucion_base)

                nueva_fila = [
                    fecha_pub, fabricante.capitalize(), cve_id, cwe, severidad_es, 
                    vector_es, tecnologia.capitalize(), poc_disponible, 
                    explotacion_activa, solucion_final, resumen_es
                ]
                
                # Sobrescribe el dato si existe, o lo añade si es nuevo
                cves_maestro[cve_id] = nueva_fila
    except Exception:
        pass 

if cves_maestro:
    print("\n[*] Ordenando los datos según reglas establecidas...")
    # Convertimos el diccionario nuevamente a una lista para ordenarla
    todas_las_filas = list(cves_maestro.values())
    todas_las_filas.sort(key=reglas_de_orden)

    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row)

    for rango in list(ws.conditional_formatting._cf_rules.keys()):
        del ws.conditional_formatting._cf_rules[rango]

    font_body = Font(name="Calibri", size=11)
    ultima_fila = 1

    print("[*] Escribiendo archivo Excel...")
    for fila_datos in todas_las_filas:
        ultima_fila += 1
        for col_idx, valor in enumerate(fila_datos, 1):
            celda = ws.cell(row=ultima_fila, column=col_idx, value=valor)
            celda.font = font_body
            celda.border = thin_border
            if col_idx in [1, 3, 4, 5, 6, 8, 9]:
                celda.alignment = Alignment(horizontal="center", vertical="center")
            else:
                celda.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[ultima_fila].height = 45 

    for col_letter in ['J', 'K']: ws.column_dimensions[col_letter].width = 50
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']: ws.column_dimensions[col_letter].width = 18

    print("[*] Aplicando Formato Condicional...")
    critica_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    critica_font = Font(color="9C0006", bold=True)
    alta_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    alta_font = Font(color="C65911", bold=True)
    media_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    media_font = Font(color="9C5700", bold=True)
    baja_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    baja_font = Font(color="006100", bold=True)
    blanco_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    blanco_font = Font(color="000000", bold=False)

    rango_severidad = f"E2:E{ws.max_row}"
    ws.conditional_formatting.add(rango_severidad, FormulaRule(formula=[f'ISNUMBER(SEARCH("Crítica", E2))'], fill=critica_fill, font=critica_font))
    ws.conditional_formatting.add(rango_severidad, FormulaRule(formula=[f'ISNUMBER(SEARCH("Alta", E2))'], fill=alta_fill, font=alta_font))
    ws.conditional_formatting.add(rango_severidad, FormulaRule(formula=[f'ISNUMBER(SEARCH("Media", E2))'], fill=media_fill, font=media_font))
    ws.conditional_formatting.add(rango_severidad, FormulaRule(formula=[f'ISNUMBER(SEARCH("Baja", E2))'], fill=baja_fill, font=baja_font))
    ws.conditional_formatting.add(rango_severidad, FormulaRule(formula=[f'ISNUMBER(SEARCH("No evaluado", E2))'], fill=blanco_fill, font=blanco_font))

    # Dashboard
    print("[*] Generando Dashboard Resumen...")
    conteo_sev, conteo_vec, conteo_fab = Counter(), Counter(), Counter()
    
    for fila in todas_las_filas:
        sev_limpia = str(fila[4]).split(" (")[0].strip()
        conteo_sev[sev_limpia] += 1
        conteo_vec[str(fila[5]).strip()] += 1
        conteo_fab[str(fila[1]).strip()] += 1

    titulo_dash = "Dashboard Resumen"
    if titulo_dash in wb.sheetnames: del wb[titulo_dash]
    ws_dash = wb.create_sheet(title=titulo_dash)
    ws_dash.views.sheetView[0].showGridLines = False

    ws_dash.append(["Severidad", "Cantidad"])
    for k, v in conteo_sev.items(): ws_dash.append([k, v])
    len_sev = len(conteo_sev)

    fila_inicio_vec = 1
    ws_dash.cell(row=fila_inicio_vec, column=5, value="Vector")
    ws_dash.cell(row=fila_inicio_vec, column=6, value="Cantidad")
    for i, (k, v) in enumerate(conteo_vec.items(), 1):
        ws_dash.cell(row=fila_inicio_vec + i, column=5, value=k)
        ws_dash.cell(row=fila_inicio_vec + i, column=6, value=v)
    len_vec = len(conteo_vec)

    fila_inicio_fab = 1
    ws_dash.cell(row=fila_inicio_fab, column=9, value="Fabricante")
    ws_dash.cell(row=fila_inicio_fab, column=10, value="Cantidad")
    for i, (k, v) in enumerate(conteo_fab.most_common(15), 1):
        ws_dash.cell(row=fila_inicio_fab + i, column=9, value=k)
        ws_dash.cell(row=fila_inicio_fab + i, column=10, value=v)
    len_fab = len(conteo_fab.most_common(15))

    grafico_sev = PieChart()
    grafico_sev.title = "Nivel de Severidad"
    grafico_sev.add_data(Reference(ws_dash, min_col=2, min_row=1, max_row=len_sev+1), titles_from_data=True)
    grafico_sev.set_categories(Reference(ws_dash, min_col=1, min_row=2, max_row=len_sev+1))
    grafico_sev.width, grafico_sev.height = 14, 10
    ws_dash.add_chart(grafico_sev, "B12")

    grafico_vec = PieChart()
    grafico_vec.title = "Vector de Ejecución"
    grafico_vec.add_data(Reference(ws_dash, min_col=6, min_row=1, max_row=len_vec+1), titles_from_data=True)
    grafico_vec.set_categories(Reference(ws_dash, min_col=5, min_row=2, max_row=len_vec+1))
    grafico_vec.width, grafico_vec.height = 14, 10
    ws_dash.add_chart(grafico_vec, "I12")

    grafico_fab = BarChart()
    grafico_fab.type, grafico_fab.style = "col", 10
    grafico_fab.title, grafico_fab.y_axis.title = "Top Fabricantes Vulnerables", "Número de CVEs"
    grafico_fab.add_data(Reference(ws_dash, min_col=10, min_row=1, max_row=len_fab+1), titles_from_data=True)
    grafico_fab.set_categories(Reference(ws_dash, min_col=9, min_row=2, max_row=len_fab+1))
    grafico_fab.legend = None
    grafico_fab.width, grafico_fab.height = 28, 12
    ws_dash.add_chart(grafico_fab, "B35")

    wb.move_sheet("Dashboard Resumen", offset=-1)
    wb.save(ARCHIVO_EXCEL)
    
    # Limpieza de txt
    open(ARCHIVO_TXT, 'w').close()
    
    print(f"[+] ¡ÉXITO! Bitácora actualizada (Inserciones y Actualizaciones).")
    print(f"[+] El archivo '{ARCHIVO_TXT}' ha sido vaciado para la próxima ejecución del Watcher.")
else:
    print("\n[-] No hay información para procesar.")